#!/usr/bin/env python3
"""
Synthetic workbooks for the failure-path tests.

Deliberately not the real training plan: these are throwaway files with known
shapes, so a test that fails points at the app rather than at somebody's data.

    python3 tests/make-fixtures.py            # writes into tests/fixtures/

Needs openpyxl (pip install openpyxl). The tests themselves need Playwright and
a static server on port 7810:

    python3 -m http.server 7810
    node tests/failure-paths.js
"""
import datetime
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'fixtures')

HEADERS = ['Week', 'Date', 'Day', 'Sport', 'Workout', 'Duration (min)',
           'Intensity', 'Purpose', 'Done', 'Actual (min)', 'Actual (km)',
           'Avg HR', 'Effort']

DAYS = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']


def monday_of_this_week():
    today = datetime.date.today()
    return today - datetime.timedelta(days=today.weekday())


def plain(path):
    """An ordinary week: something on most days, one rest day, one blank day."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Weekly Schedules'
    ws.append(HEADERS)

    monday = monday_of_this_week()
    week = [
        (2, 'Swim', 'Easy technique: 4x50 drill, then 6x50 smooth', 30, 'Z1-Z2', 'Feel for the water'),
        (2, 'Mobility', 'Hips and T-spine', 15, '-', 'Keep the hips honest'),
        (3, 'Run', 'Easy run, conversational the whole way', 30, 'Z2', 'Aerobic base'),
        (4, 'Rest', 'REST DAY - full day off', None, '-', 'Adaptation happens at rest'),
        (5, 'Bike', 'Steady aerobic ride, flat to rolling', 60, 'Z2', 'Time in the saddle'),
        (6, 'Run', 'Easy run + strides', 35, 'Z2', 'Turnover'),
        (6, 'Strength', 'Squat, hinge, press, row', 30, 'RPE 7', 'Durability'),
    ]
    next_week = [
        (7, 'Swim', 'Technique: 8x50 drills + 6x100 easy', 35, 'Z1-Z2', 'Stroke before volume'),
        (8, 'Bike', 'Easy spin, cadence 90+', 45, 'Z1', 'Legs turning, nothing more'),
        (9, 'Run', 'Easy run + 4x20 s strides', 35, 'Z2', 'Turnover'),
        (11, 'Rest', 'REST DAY - full day off', None, '-', 'The weekly anchor'),
        (12, 'Bike', 'Steady aerobic ride', 65, 'Z2', 'Long, but not hard'),
        (13, 'Run', 'Easy run, flat route', 40, 'Z2', 'Aerobic base'),
    ]

    # Two weeks, so "next week" is a real week rather than an empty one.
    for offset, sport, workout, minutes, zone, purpose in week + next_week:
        date = monday + datetime.timedelta(days=offset)
        ws.append([1 if offset < 7 else 2, date.isoformat(), DAYS[date.weekday()], sport,
                   workout, minutes, zone, purpose])

    last = ws.max_row + 1
    ws.cell(last, 5, 'Weekly total')
    ws.cell(last, 6, '=SUM(F2:F%d)' % (last - 1))
    wb.save(path)


def hostile_text(path):
    """Every cell that reaches the screen, carrying something that should not."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Weekly Schedules'
    ws.append(HEADERS)

    monday = monday_of_this_week()
    rows = [
        ('Swim', '<img src=x onerror="window.hacked=1">', 30, 'injection attempt'),
        ('Run', '</p><script>window.hacked=2</script>', 40, '"quotes" & <ampersands>'),
        ('Bike', "onclick=alert(1) '\" ><", 45, 'attribute break-out'),
        ('Strength', 'A' * 4000, 35, 'a title with no end to it'),
    ]
    for index, (sport, workout, minutes, purpose) in enumerate(rows):
        date = monday + datetime.timedelta(days=index)
        ws.append([1, date.isoformat(), DAYS[date.weekday()], sport,
                   workout, minutes, 'Z2', purpose])
    wb.save(path)


def column_inserted(path):
    """The same plan with one column pushed in on the left.

    Every heading to its right moves one column over, which is exactly what a
    saved layout cannot see for itself."""
    wb = openpyxl.load_workbook(os.path.join(OUT, 'plain.xlsx'))
    ws = wb['Weekly Schedules']
    ws.insert_cols(2)
    ws.cell(1, 2, 'Notes to self')
    ws.cell(2, 2, 'inserted in Excel')
    wb.save(path)


def foreign_extras(path):
    """Somebody else's sheet already sitting on the name the app wants."""
    wb = openpyxl.load_workbook(os.path.join(OUT, 'plain.xlsx'))
    ws = wb.create_sheet('Extras')
    ws.append(['Race entry', 'Cost', 'Paid?'])
    ws.append(['Kalmar 2027', 4200, 'Yes'])
    ws.append(['Jonkoping 70.3', 1900, 'No'])
    wb.save(path)


def history(path):
    """
    Twelve weeks already behind us, with a deliberate shape to find:

      - Thursday is the day that slips (most Thursdays missed or ignored)
      - Swim is the sport that runs behind
      - a run of completed sessions at the end, for the streak
      - some sessions left unanswered, which must not read as completed

    Rest days are included precisely because they must NOT be counted: they
    cannot be kept or missed, and counting them would flatter every figure.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Weekly Schedules'
    ws.append(HEADERS)

    monday = monday_of_this_week()
    start = monday - datetime.timedelta(weeks=12)

    # (weekday, sport, minutes)
    template = [
        (0, 'Swim', 45),
        (1, 'Bike', 60),
        (2, 'Run', 40),
        (3, 'Swim', 45),        # Thursday: the one that gets skipped
        (4, 'Rest', None),
        (5, 'Bike', 90),
        (6, 'Strength', 30),
    ]

    for week in range(12):
        for weekday, sport, minutes in template:
            date = start + datetime.timedelta(weeks=week, days=weekday)
            row = [week + 1, date, DAYS[weekday], sport,
                   f'{sport} session', minutes, 'Z2', 'Base']

            if sport == 'Rest':
                ws.append(row + ['', None, None, None, None])
                continue

            done = ''
            actual = None
            # Thursday slips; swim slips generally; the last fortnight is clean.
            if week >= 10:
                done, actual = '\u2713', minutes
            elif weekday == 3:
                done = 'Missed' if week % 2 else ''      # missed, or never answered
            elif sport == 'Swim' and week % 3 == 0:
                done = 'Missed'
            else:
                done, actual = '\u2713', minutes

            ws.append(row + [done, actual, None, None, None])

    wb.save(path)


def row_inserted(path, source):
    """
    The same week as plain.xlsx with one extra row pushed in above the data —
    a phase banner, a note, anything somebody types in Excel on a Sunday.

    Every session below it moves down one row, which changes the identity the
    app derives from sheet+row. Anything the app remembers *by* that identity
    now points at the wrong session.
    """
    wb = openpyxl.load_workbook(source)
    ws = wb['Weekly Schedules']
    ws.insert_rows(2)
    ws.cell(row=2, column=1, value='')
    ws.cell(row=2, column=2, value='--- a note somebody added ---')
    wb.save(path)


def paced(path):
    """
    A week with a swim, a bike and a run today, and a single shared
    "Avg Pace/Pwr" column — the shape a real triathlon plan uses, where one
    column has to answer for three sports that measure themselves differently.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Weekly Schedules'
    ws.append(['Week', 'Date', 'Day', 'Sport', 'Workout', 'Duration (min)',
               'Intensity', 'Purpose', 'Done', 'Actual (min)', 'Actual (km)',
               'Avg Pace/Pwr', 'Avg HR', 'Effort', 'Notes'])

    today = datetime.date.today()
    for sport, minutes, what in [
        ('Swim', 45, 'Technique: 8x50 drills'),
        ('Bike', 90, 'Steady aerobic ride'),
        ('Run', 40, 'Easy run, conversational'),
    ]:
        ws.append([1, today, DAYS[today.weekday()], sport, what, minutes,
                   'Z2', 'Base', '', None, None, None, None, None, None])

    wb.save(path)


def everyday(path):
    """
    Two sessions on every day of this week and next, no rest days, no blanks.

    Exists for the tests that drive *today's* card — logging it, marking it
    missed, watching the buttons collapse. plain.xlsx starts its week on
    Wednesday and rests on Friday, so any test that needed "something to log
    today" was green three days a week and red the other four, which is how a
    suite quietly trains people to rerun it until it passes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Weekly Schedules'
    ws.append(HEADERS)

    monday = monday_of_this_week()
    sports = [('Run', 35, 'Z2'), ('Swim', 30, 'Z1-Z2')]
    for offset in range(14):
        date = monday + datetime.timedelta(days=offset)
        for sport, minutes, zone in sports:
            ws.append([1 if offset < 7 else 2, date.isoformat(), DAYS[date.weekday()],
                       sport, sport + ' session', minutes, zone, 'Base'])
    wb.save(path)


def broken(path_rubbish, path_truncated, source):
    with open(path_rubbish, 'wb') as fh:
        fh.write(b'this is not a zip, it is a sentence')
    with open(source, 'rb') as fh:
        good = fh.read()
    with open(path_truncated, 'wb') as fh:
        fh.write(good[:int(len(good) * 0.6)])


if __name__ == '__main__':
    os.makedirs(OUT, exist_ok=True)
    plain(os.path.join(OUT, 'plain.xlsx'))
    hostile_text(os.path.join(OUT, 'nasty.xlsx'))
    column_inserted(os.path.join(OUT, 'column-inserted.xlsx'))
    foreign_extras(os.path.join(OUT, 'foreign-extras.xlsx'))
    history(os.path.join(OUT, 'history.xlsx'))
    paced(os.path.join(OUT, 'paced.xlsx'))
    everyday(os.path.join(OUT, 'everyday.xlsx'))
    row_inserted(os.path.join(OUT, 'row-inserted.xlsx'),
                 os.path.join(OUT, 'plain.xlsx'))
    broken(os.path.join(OUT, 'rubbish.xlsx'),
           os.path.join(OUT, 'truncated.xlsx'),
           os.path.join(OUT, 'plain.xlsx'))
    print('fixtures written to', OUT)
